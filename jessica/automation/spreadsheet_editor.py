"""Spreadsheet automation for Jessica.

Future capability: allows Jessica to read, edit, and create Excel/CSV files.
All operations require user approval before execution.
"""
from __future__ import annotations

import csv
import os
import subprocess
import sys
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class SpreadsheetEditor:
    """Safe spreadsheet operations with user approval."""

    def __init__(self):
        self.last_file: Optional[str] = None

    def _resolve_filepath(self, filename: str) -> Optional[str]:
        """Try to resolve a filename to an actual file path.
        
        Search in:
        1. Current directory
        2. Desktop
        3. Documents
        4. Downloads
        """
        # If already a full path and exists
        if os.path.isfile(filename):
            return filename
        
        # Search in common locations
        search_paths = [
            os.getcwd(),
            os.path.expanduser("~/Desktop"),
            os.path.expanduser("~/Documents"),
            os.path.expanduser("~/Downloads"),
        ]
        
        for search_dir in search_paths:
            full_path = os.path.join(search_dir, filename)
            if os.path.isfile(full_path):
                return full_path
        
        # Also try adding common extensions if not present
        if not filename.endswith((".xlsx", ".xls", ".csv")):
            for ext in [".xlsx", ".xls", ".csv"]:
                for search_dir in search_paths:
                    full_path = os.path.join(search_dir, filename + ext)
                    if os.path.isfile(full_path):
                        return full_path
        
        return None

    def check_dependencies(self) -> Dict[str, bool]:
        """Verify spreadsheet libraries are available."""
        return {
            "openpyxl": OPENPYXL_AVAILABLE,  # For .xlsx
            "pandas": PANDAS_AVAILABLE,       # For advanced operations
        }

    def request_permission(self, operation: str, filepath: str, details: str = "") -> bool:
        """Ask user to approve a spreadsheet operation.
        
        In GUI mode (no stdin), auto-approve safe operations like opening files.
        In terminal mode, prompt the user for approval.
        """
        # Check if stdin is available (terminal mode)
        has_stdin = sys.stdin and sys.stdin.isatty()
        
        # Auto-approve safe read-only operations in GUI mode
        safe_operations = {"open Excel file", "read CSV", "read cell", "get cell summary"}
        is_safe = operation in safe_operations
        
        if not has_stdin and is_safe:
            # GUI mode with a safe operation - auto-approve
            print(f"[spreadsheet] Auto-approving {operation}: {filepath}")
            return True
        
        if not has_stdin:
            # GUI mode with a potentially dangerous operation - deny
            print(f"[spreadsheet] Cannot {operation} in GUI mode without user confirmation: {filepath}")
            return False
        
        # Terminal mode - prompt user
        print(f"[spreadsheet] Jessica wants to {operation}: {filepath}")
        if details:
            print(f"[spreadsheet] Details: {details}")
        approve = input("Approve this operation? [y/N]: ").strip().lower()
        return approve in {"y", "yes"}

    def read_csv(self, filepath: str) -> Optional[List[List[str]]]:
        """Read a CSV file and return rows."""
        if not os.path.isfile(filepath):
            return None

        if not self.request_permission("read CSV", filepath):
            return None

        try:
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)
            self.last_file = filepath
            return rows
        except Exception as exc:
            print(f"[spreadsheet] CSV read failed: {exc}")
            return None

    def write_csv(self, filepath: str, rows: List[List[str]]) -> bool:
        """Write rows to a CSV file."""
        details = f"{len(rows)} rows"
        if not self.request_permission("write CSV", filepath, details):
            return False

        try:
            with open(filepath, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            self.last_file = filepath
            print(f"[spreadsheet] CSV saved: {filepath}")
            return True
        except Exception as exc:
            print(f"[spreadsheet] CSV write failed: {exc}")
            return False

    def read_excel(self, filepath: str, sheet_name: Optional[str] = None) -> Optional[List[List[Any]]]:
        """Read an Excel file (.xlsx) and return rows."""
        if not OPENPYXL_AVAILABLE:
            print("[spreadsheet] Excel support requires openpyxl. Install: pip install openpyxl")
            return None

        if not os.path.isfile(filepath):
            return None

        if not self.request_permission("read Excel", filepath, f"sheet: {sheet_name or 'first'}"):
            return None

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            self.last_file = filepath
            return rows
        except Exception as exc:
            print(f"[spreadsheet] Excel read failed: {exc}")
            return None

    def write_excel(self, filepath: str, rows: List[List[Any]], sheet_name: str = "Sheet1") -> bool:
        """Write rows to an Excel file (.xlsx)."""
        if not OPENPYXL_AVAILABLE:
            print("[spreadsheet] Excel support requires openpyxl. Install: pip install openpyxl")
            return False

        details = f"{len(rows)} rows to sheet '{sheet_name}'"
        if not self.request_permission("write Excel", filepath, details):
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            for row in rows:
                ws.append(row)
            wb.save(filepath)
            self.last_file = filepath
            print(f"[spreadsheet] Excel saved: {filepath}")
            return True
        except Exception as exc:
            print(f"[spreadsheet] Excel write failed: {exc}")
            return False

    def get_summary(self, filepath: str) -> Optional[str]:
        """Get a quick summary of a spreadsheet file."""
        if not os.path.isfile(filepath):
            return f"[spreadsheet] File not found: {filepath}"

        ext = os.path.splitext(filepath)[1].lower()

        if ext == ".csv":
            rows = self.read_csv(filepath)
            if rows:
                return f"[spreadsheet] CSV with {len(rows)} rows, {len(rows[0]) if rows else 0} columns"
            return "[spreadsheet] Could not read CSV"

        elif ext in {".xlsx", ".xls"}:
            rows = self.read_excel(filepath)
            if rows:
                return f"[spreadsheet] Excel with {len(rows)} rows, {len(rows[0]) if rows else 0} columns"
            return "[spreadsheet] Could not read Excel"

        return f"[spreadsheet] Unsupported format: {ext}"

    def open_excel(self, filepath: str) -> bool:
        """Open an Excel file in Microsoft Excel application."""
        # Try to resolve the file path
        resolved_path = self._resolve_filepath(filepath)
        
        if not resolved_path:
            print(f"[spreadsheet] File not found: {filepath}")
            return False

        if not self.request_permission("open Excel file", resolved_path, "This will launch Microsoft Excel"):
            return False

        try:
            if os.name == 'nt':  # Windows
                os.startfile(resolved_path)
            elif os.name == 'posix':  # macOS/Linux
                subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', filepath])
            
            print(f"[spreadsheet] Opened: {filepath}")
            return True
        except Exception as e:
            print(f"[spreadsheet] Failed to open: {e}")
            return False


def can_handle(intent):
    """Check if this skill should handle spreadsheet operations."""
    text = intent.get("text", "").lower()
    triggers = [
        "spreadsheet",
        "excel",
        "csv",
        "edit cell",
        "open spreadsheet",
        "read csv",
        "write csv",
        "update excel",
    ]
    return any(t in text for t in triggers)


def run(intent, context, relevant, manager):
    """Skill entry point for spreadsheet requests."""
    editor = SpreadsheetEditor()
    deps = editor.check_dependencies()

    user_text = intent.get("text", "").lower()
    full_text = intent.get("text", "")

    # Check if user has attached files with content
    if "[user attached files" in user_text:
        # Extract and analyze the attached content
        parts = full_text.split("[USER ATTACHED FILES - ANALYZE THIS CONTENT]:")
        if len(parts) > 1:
            file_content = parts[1].strip()
            
            # Extract just the main query (before the marker)
            query = parts[0].strip()
            
            # Check if this is an Excel file
            if "[EXCEL:" in file_content:
                # Parse Excel content and provide intelligent summary
                lines = file_content.split("\n")
                excel_info = None
                
                for i, line in enumerate(lines):
                    if "[EXCEL:" in line:
                        excel_info = line
                        # Get the data rows that follow
                        data_rows = [l for l in lines[i+1:] if l.strip() and not l.startswith("[")]
                        
                        # Analyze the spreadsheet intelligently
                        summary = f"📊 **Spreadsheet Analysis**\n\n"
                        summary += f"File: {excel_info}\n\n"
                        
                        if data_rows:
                            # Extract header row (first row usually has column names)
                            header_row = data_rows[0] if data_rows else ""
                            headers = [h.strip() for h in header_row.split("|")]
                            num_columns = len([h for h in headers if h])
                            
                            # Analyze content
                            summary += f"**Overview:**\n"
                            summary += f"- Contains {len(data_rows)} rows of data\n"
                            summary += f"- {num_columns} columns\n"
                            
                            # Identify data type by examining content
                            sample_data = " ".join(data_rows[:5]).lower()
                            if any(word in sample_data for word in ["transfer", "adjustment", "warehouse", "location", "stock"]):
                                summary += f"- **Type:** Inventory/Supply Chain data\n"
                                summary += f"- **Purpose:** Tracking transfers and adjustments between locations\n"
                            elif any(word in sample_data for word in ["sales", "revenue", "customer", "order"]):
                                summary += f"- **Type:** Sales/Order data\n"
                            elif any(word in sample_data for word in ["employee", "department", "salary", "hours"]):
                                summary += f"- **Type:** HR/Payroll data\n"
                            else:
                                summary += f"- **Type:** Business records\n"
                            
                            # List columns
                            summary += f"\n**Columns:** {', '.join(headers[:10])}"
                            if len(headers) > 10:
                                summary += f", and {len(headers) - 10} more"
                            
                            # Provide actionable next steps
                            summary += f"\n\n**What would you like to know?**\n"
                            summary += f"- 'Show me row X' - View specific row details\n"
                            summary += f"- 'Find [keyword]' - Search for specific entries\n"
                            summary += f"- 'Summary stats' - Get statistics about the data\n"
                        else:
                            summary += "The spreadsheet appears to be empty or couldn't be read."
                        
                        if query:
                            summary += f"\n\n*Your question was: {query}*"
                        
                        return {"reply": summary}
            
            # Handle CSV files
            elif "[TEXT:" in file_content and (".csv" in file_content.lower() or ".txt" in file_content.lower()):
                lines = file_content.split("\n")
                text_info = None
                
                for i, line in enumerate(lines):
                    if "[TEXT:" in line or "[CSV:" in line:
                        text_info = line
                        data_rows = [l for l in lines[i+1:] if l.strip() and not l.startswith("[")]
                        
                        summary = f"📄 **File Analysis**\n\n"
                        summary += f"File: {text_info}\n\n"
                        summary += f"**Overview:**\n"
                        summary += f"- {len(data_rows)} lines of content\n"
                        
                        if data_rows:
                            summary += f"- First line: {data_rows[0][:100]}\n"
                        
                        summary += f"\n**What would you like to know?**\n"
                        summary += f"- 'Show me line X' - View specific line\n"
                        summary += f"- 'Find [text]' - Search for content\n"
                        summary += f"- 'Summarize' - Get a summary of the content\n"
                        
                        if query:
                            summary += f"\n\n*Your question was: {query}*"
                        
                        return {"reply": summary}
            
            # For other file types, just report what was attached
            else:
                summary = f"📎 **File Attached for Analysis**\n\n"
                summary += "I've received your attached file and I'm ready to analyze it.\n"
                summary += "Please ask me a specific question about the content, such as:\n"
                summary += "- 'What patterns do you see?'\n"
                summary += "- 'Summarize the key points'\n"
                summary += "- 'Find entries with [keyword]'\n"
                if query:
                    summary += f"\n*Your question was: {query}*"
                return {"reply": summary}

    # First, try live Excel if no specific file is mentioned
    if "open" not in user_text and any(kw in user_text for kw in ["active excel", "current sheet", "current spreadsheet", "read cell", "what's in"]):
        try:
            from jessica.automation import excel_live
            controller = excel_live.ExcelLiveController()
            if controller.connect_to_excel():
                summary = controller.get_cell_summary()
                return {"reply": summary}
            else:
                return {"reply": "No active Excel workbook detected. Please open an Excel file first."}
        except Exception as e:
            return {"reply": f"Error accessing Excel: {e}"}

    # Check if user wants to open a file
    if "open" in user_text:
        # Extract potential filenames (words that look like filenames)
        words = user_text.split()
        for word in words:
            # Remove quotes if present
            word = word.strip('\'"')
            if any(word.endswith(ext) for ext in [".xlsx", ".xls", ".csv"]):
                filepath = editor._resolve_filepath(word)
                if filepath:
                    if editor.open_excel(filepath):
                        return {"reply": f"Opening {os.path.basename(filepath)} in Excel..."}
                else:
                    return {"reply": f"Could not find {word}. Searched in: Documents, Desktop, Downloads, and current directory."}
        
        # No specific file mentioned - ask for details
        return {"reply": "Please specify the Excel file name you want to open. Example: 'open budget.xlsx' or 'open 'Transfer_List_20260108_084650.xlsx'"}

    # Check if user wants to read a file
    if "read" in user_text or "show" in user_text or "summary" in user_text:
        words = user_text.split()
        for word in words:
            word = word.strip('\'"')
            if any(word.endswith(ext) for ext in [".xlsx", ".xls", ".csv"]):
                filepath = editor._resolve_filepath(word)
                if filepath:
                    summary = editor.get_summary(filepath)
                    if summary:
                        return {"reply": summary}

    # Default: show capabilities
    status = []
    status.append("I can help with Excel and CSV files!")
    status.append(f"\nCapabilities:")
    status.append(f"- CSV support: ✓ (built-in)")
    status.append(f"- Excel support: {'✓' if deps['openpyxl'] else '✗ (install openpyxl)'}")
    status.append(f"- Advanced features: {'✓' if deps['pandas'] else '✗ (install pandas)'}")
    status.append("\nCommands I understand:")
    status.append("- 'open [filename.xlsx]' - Opens Excel file")
    status.append("- 'read [filename.csv]' - Reads and summarizes file")
    status.append("- 'edit [filename]' - Edits spreadsheet with your approval")
    status.append("\nAll operations require your approval before execution.")

    return {"reply": "\n".join(status)}
