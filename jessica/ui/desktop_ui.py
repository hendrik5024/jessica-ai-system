"""CustomTkinter-based desktop UI for Jessica.

Modern, dark-themed interface with:
- Chat window with streaming support
- Voice input button (push-to-talk)
- Text-to-speech toggle
- Settings panel
"""
from __future__ import annotations

import threading
import time
from typing import Any, Optional
from pathlib import Path

try:
    import customtkinter as ctk
    CTK_AVAILABLE = True
except ImportError:
    CTK_AVAILABLE = False
    import tkinter as tk
    import tkinter.ttk as ttk
    from tkinter import filedialog

from jessica.jessica_core import CognitiveManager
from jessica.tts.tts_adapter import speak
from jessica.stt.listener import VoiceListener


class JessicaDesktopUI:
    """CustomTkinter-based desktop interface for Jessica."""

    def __init__(self, manager: Optional[CognitiveManager] = None):
        self.manager = manager or CognitiveManager()
        self.voice_listener: Optional[VoiceListener] = None
        self.tts_enabled = True
        self.is_recording = False
        self.approval_response = None
        self.attached_files: list[Path] = []  # Store attached files
        
        # Initialize window
        if CTK_AVAILABLE:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.root = ctk.CTk()
        else:
            self.root = tk.Tk()
        
        self.root.title("Jessica - Offline AI Assistant")
        self.root.geometry("900x700")
        
        self._build_ui()
        self._start_voice_listener()

    def _build_ui(self):
        """Build the user interface."""
        # Main container
        if CTK_AVAILABLE:
            main_frame = ctk.CTkFrame(self.root)
        else:
            main_frame = tk.Frame(self.root, bg="#2b2b2b")
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Title
        if CTK_AVAILABLE:
            title = ctk.CTkLabel(
                main_frame,
                text="Jessica AI Assistant",
                font=ctk.CTkFont(size=24, weight="bold")
            )
        else:
            title = tk.Label(
                main_frame,
                text="Jessica AI Assistant",
                font=("Arial", 24, "bold"),
                bg="#2b2b2b",
                fg="white"
            )
        title.pack(pady=(0, 10))
        
        # Chat display
        if CTK_AVAILABLE:
            self.chat_display = ctk.CTkTextbox(
                main_frame,
                width=850,
                height=450,
                font=ctk.CTkFont(size=13),
                wrap="word"
            )
        else:
            self.chat_display = tk.Text(
                main_frame,
                width=100,
                height=25,
                font=("Consolas", 11),
                bg="#1e1e1e",
                fg="white",
                wrap="word"
            )
        self.chat_display.pack(pady=10, padx=10, fill="both", expand=True)
        self.chat_display.configure(state="disabled")
        
        # Input frame
        if CTK_AVAILABLE:
            input_frame = ctk.CTkFrame(main_frame)
        else:
            input_frame = tk.Frame(main_frame, bg="#2b2b2b")
        input_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        # Text input
        if CTK_AVAILABLE:
            self.text_input = ctk.CTkEntry(
                input_frame,
                placeholder_text="Type your message...",
                font=ctk.CTkFont(size=13),
                height=40
            )
        else:
            self.text_input = tk.Entry(
                input_frame,
                font=("Arial", 12),
                bg="#1e1e1e",
                fg="white"
            )
        self.text_input.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.text_input.bind("<Return>", lambda e: self._send_message())
        
        # Enable paste functionality
        self.text_input.bind("<Control-v>", lambda e: self._paste_text())
        self.text_input.bind("<Control-V>", lambda e: self._paste_text())
        
        # Send button
        if CTK_AVAILABLE:
            send_btn = ctk.CTkButton(
                input_frame,
                text="Send",
                command=self._send_message,
                width=80,
                height=40
            )
        else:
            send_btn = tk.Button(
                input_frame,
                text="Send",
                command=self._send_message,
                bg="#0066cc",
                fg="white",
                font=("Arial", 11)
            )
        send_btn.pack(side="left", padx=2)
        
        # Voice button
        if CTK_AVAILABLE:
            self.voice_btn = ctk.CTkButton(
                input_frame,
                text="🎤 Voice",
                command=self._toggle_voice_input,
                width=80,
                height=40,
                fg_color="green"
            )
        else:
            self.voice_btn = tk.Button(
                input_frame,
                text="🎤 Voice",
                command=self._toggle_voice_input,
                bg="green",
                fg="white",
                font=("Arial", 11)
            )
        self.voice_btn.pack(side="left", padx=2)
        
        # Paperclip button (attach files/images)
        if CTK_AVAILABLE:
            attach_btn = ctk.CTkButton(
                input_frame,
                text="📎 Attach",
                command=self._attach_files,
                width=80,
                height=40,
                fg_color="orange"
            )
        else:
            attach_btn = tk.Button(
                input_frame,
                text="📎 Attach",
                command=self._attach_files,
                bg="#ff9900",
                fg="white",
                font=("Arial", 11)
            )
        attach_btn.pack(side="left", padx=2)
        
        # Control frame
        if CTK_AVAILABLE:
            control_frame = ctk.CTkFrame(main_frame)
        else:
            control_frame = tk.Frame(main_frame, bg="#2b2b2b")
        control_frame.pack(fill="x", padx=10)
        
        # TTS toggle
        if CTK_AVAILABLE:
            self.tts_toggle = ctk.CTkSwitch(
                control_frame,
                text="Text-to-Speech",
                command=self._toggle_tts,
                onvalue=True,
                offvalue=False
            )
            self.tts_toggle.select()
        else:
            self.tts_var = tk.BooleanVar(value=True)
            self.tts_toggle = tk.Checkbutton(
                control_frame,
                text="Text-to-Speech",
                variable=self.tts_var,
                command=self._toggle_tts,
                bg="#2b2b2b",
                fg="white",
                selectcolor="#1e1e1e"
            )
        self.tts_toggle.pack(side="left", padx=10)
        
        # TTS self-test button
        if CTK_AVAILABLE:
            tts_test_btn = ctk.CTkButton(
                control_frame,
                text="TTS Test",
                command=self._tts_self_test,
                width=90,
                height=32
            )
        else:
            tts_test_btn = tk.Button(
                control_frame,
                text="TTS Test",
                command=self._tts_self_test,
                bg="#444444",
                fg="white",
                font=("Arial", 10)
            )
        tts_test_btn.pack(side="left", padx=6)

        # Free Access toggle (auto-approve permissions)
        if CTK_AVAILABLE:
            self.free_access_toggle = ctk.CTkSwitch(
                control_frame,
                text="Free Access",
                command=self._toggle_free_access,
                onvalue=True,
                offvalue=False
            )
        else:
            self.free_access_var = tk.BooleanVar(value=False)
            self.free_access_toggle = tk.Checkbutton(
                control_frame,
                text="Free Access",
                variable=self.free_access_var,
                command=self._toggle_free_access,
                bg="#2b2b2b",
                fg="white",
                selectcolor="#1e1e1e"
            )
        self.free_access_toggle.pack(side="left", padx=10)
        
        # Privacy Settings Button
        if CTK_AVAILABLE:
            privacy_btn = ctk.CTkButton(
                control_frame,
                text="⚙️ Privacy",
                command=self._show_privacy_settings,
                width=90,
                height=32
            )
        else:
            privacy_btn = tk.Button(
                control_frame,
                text="⚙️ Privacy",
                command=self._show_privacy_settings,
                bg="#444444",
                fg="white",
                font=("Arial", 10)
            )
        privacy_btn.pack(side="left", padx=6)
        
        # Status label
        if CTK_AVAILABLE:
            self.status_label = ctk.CTkLabel(
                control_frame,
                text="Ready. Say 'Jessica, <command>' while holding Ctrl+Shift+Space",
                font=ctk.CTkFont(size=11)
            )
        else:
            self.status_label = tk.Label(
                control_frame,
                text="Ready. Say 'Jessica, <command>' while holding Ctrl+Shift+Space",
                bg="#2b2b2b",
                fg="gray",
                font=("Arial", 10)
            )
        self.status_label.pack(side="left", padx=10)
        
        # Add welcome message
        self._append_message("System", "Jessica AI Assistant ready. Type or use voice commands.", "info")

    def _append_message(self, sender: str, message: str, msg_type: str = "normal"):
        """Append a message to the chat display."""
        self.chat_display.configure(state="normal")
        
        if msg_type == "info":
            prefix = "ℹ️ "
        elif msg_type == "error":
            prefix = "❌ "
        elif sender == "You":
            prefix = "👤 You: "
        else:
            prefix = "🤖 Jessica: "
        
        self.chat_display.insert("end", f"{prefix}{message}\n\n")
        self.chat_display.see("end")
        self.chat_display.configure(state="disabled")

    def _tts_self_test(self):
        """Run a two-phrase TTS self-test to verify audio."""
        def run_test():
            try:
                # Sync TTS state from toggle
                if CTK_AVAILABLE:
                    self.tts_enabled = self.tts_toggle.get()
                else:
                    self.tts_enabled = self.tts_var.get()
                
                if not self.tts_enabled:
                    self._append_message("System", "TTS is disabled. Enable the toggle and try again.", "error")
                    return
                
                self._append_message("System", "Running TTS self-test (2 phrases)...", "info")
                # Block until each phrase finishes, but run through the single TTS worker.
                speak("Test one. This is Jessica speaking.", async_mode=False)
                time.sleep(0.2)
                speak("Test two. Speaking again.", async_mode=False)
                self._append_message("System", "TTS self-test completed.", "info")
            except Exception as e:
                self._append_message("System", f"TTS self-test error: {e}", "error")
        threading.Thread(target=run_test, daemon=True).start()

    def _send_message(self):
        """Send text message to Jessica."""
        text = self.text_input.get().strip()
        if not text:
            return
        
        self.text_input.delete(0, "end")
        self._append_message("You", text)
        
        # Process in background thread
        threading.Thread(target=self._process_message, args=(text,), daemon=True).start()

    def _paste_text(self):
        """Handle paste from clipboard."""
        try:
            clipboard_text = self.root.clipboard_get()
            # Get current cursor position
            cursor_pos = self.text_input.index("insert") if hasattr(self.text_input, "index") else len(self.text_input.get())
            # Insert clipboard text at cursor
            self.text_input.insert(cursor_pos, clipboard_text)
            return "break"  # Prevent default paste behavior
        except Exception:
            pass  # Clipboard empty or unavailable

    def _process_message(self, text: str, attached_files: list = None):
        """Process message and get response, including attached files."""
        try:
            # Inject approval callback for vision features
            self._inject_approval_callback()
            
            # If files are attached, build a context message for Jessica
            if attached_files:
                file_context = self._extract_file_contents(attached_files)
                if file_context:
                    # Make it clear this is about analyzing the attached files
                    prompt = text if text else "Analyze the attached file(s) and tell me what you see."
                    text = f"{prompt}\n\n[USER ATTACHED FILES - ANALYZE THIS CONTENT]:\n{file_context}"
            
            response = self.manager.handle_input(text)
            reply = self._format_result(response)
            
            self.root.after(0, lambda: self._append_message("Jessica", reply))
            
            # Update TTS state from widget before checking
            if CTK_AVAILABLE:
                self.tts_enabled = self.tts_toggle.get()
            else:
                self.tts_enabled = self.tts_var.get()
            
            if self.tts_enabled:
                # Queue speech; do not block UI / avoid pyttsx3 calls from random threads.
                speak(reply, async_mode=True)
        
        except Exception as e:
            import traceback
            error_msg = f"Error: {str(e)}\n{traceback.format_exc()}"
            print(f"[UI] Error processing message: {error_msg}")
            self.root.after(0, lambda: self._append_message("System", f"Error: {str(e)}", "error"))

    def _format_result(self, resp: Any) -> str:
        """Format response for display, cleaning up internal tags."""
        import re
        
        if isinstance(resp, dict):
            # Handle skill results with 'result' key
            result = resp.get("result")
            
            if isinstance(result, dict):
                # Check for different result formats from various skills
                
                # App launcher skill format
                if "status" in result and "app" in result:
                    message = result.get("message") or f"Launching {result.get('app')}..."
                    return message
                
                # Standard formats
                if "message" in result:
                    message = result["message"]
                elif "reply" in result:
                    message = result["reply"]
                elif "output" in result:
                    message = result["output"]
                elif "error" in result:
                    return f"Error: {result['error']}"
                else:
                    message = str(result)
            
            # Handle string results
            elif isinstance(result, str):
                message = result
            
            # Try direct fields on response dict
            elif "message" in resp:
                message = resp["message"]
            elif "reply" in resp:
                message = resp["reply"]
            elif "output" in resp:
                message = resp["output"]
            elif "error" in resp:
                return f"Error: {resp['error']}"
            
            # If result exists but is None, check resp-level fields
            elif result is not None:
                message = str(result)
            else:
                return ""
        else:
            message = str(resp)
        
        # Clean up internal reasoning tags and format nicely
        # First check for combined reasoning + answer format
        reasoning_match = re.search(r'\[REASONING\](.*?)\[/REASONING\]\s*\[ANSWER\](.*?)\[/ANSWER\]', message, re.DOTALL | re.IGNORECASE)
        if reasoning_match:
            reasoning = reasoning_match.group(1).strip()
            answer = reasoning_match.group(2).strip()
            # Format nicely with clear separation
            return f"**Reasoning:**\n{reasoning}\n\n**Answer:**\n{answer}"
        
        # Check for reasoning without answer tags
        reasoning_only = re.search(r'\[REASONING\](.*?)\[/REASONING\](.*)', message, re.DOTALL | re.IGNORECASE)
        if reasoning_only:
            reasoning = reasoning_only.group(1).strip()
            answer = reasoning_only.group(2).strip()
            if answer:
                return f"**Reasoning:**\n{reasoning}\n\n{answer}"
            else:
                return reasoning
        
        # Check for answer-only tags
        answer_match = re.search(r'\[ANSWER\](.*?)\[/ANSWER\]', message, re.DOTALL | re.IGNORECASE)
        if answer_match:
            return answer_match.group(1).strip()
        
        # Remove orphaned tags if any
        message = re.sub(r'\[/?REASONING\]', '', message, flags=re.IGNORECASE)
        message = re.sub(r'\[/?ANSWER\]', '', message, flags=re.IGNORECASE)
        
        return message.strip()

    def _toggle_voice_input(self):
        """Toggle voice input recording."""
        self._append_message("System", "Voice input active. Say 'Jessica, <command>' while holding Ctrl+Shift+Space", "info")

    def _inject_approval_callback(self):
        """Inject GUI approval callback into screen monitor."""
        try:
            import jessica.vision.screen_monitor as screen_monitor
            if not hasattr(screen_monitor.ScreenMonitor, '_ui_callback_injected'):
                original_describe = screen_monitor.ScreenMonitor.describe_screen
                ui_instance = self
                
                def patched_describe(monitor_self, context="", approval_callback=None):
                    return original_describe(monitor_self, context, ui_instance._request_approval)
                
                screen_monitor.ScreenMonitor.describe_screen = patched_describe
                screen_monitor.ScreenMonitor._ui_callback_injected = True
        except Exception:
            pass  # Gracefully handle if screen_monitor not available
    
    def _request_approval(self, message: str) -> bool:
        """Request user approval via popup dialog."""
        # If free access enabled, auto-approve
        try:
            free = self.free_access_toggle.get() if CTK_AVAILABLE else self.free_access_var.get()
            if free:
                return True
        except Exception:
            pass
        self.approval_response = None
        self.root.after(0, lambda: self._show_approval_dialog(message))
        
        # Wait for user response (with timeout)
        timeout = 30  # 30 seconds
        for _ in range(timeout * 10):
            if self.approval_response is not None:
                return self.approval_response
            time.sleep(0.1)
        return False  # Timeout, deny by default
    
    def _show_approval_dialog(self, message: str):
        """Show approval request in popup dialog."""
        if CTK_AVAILABLE:
            import customtkinter as ctk
            dialog = ctk.CTkToplevel(self.root)
            dialog.title("Permission Required")
            dialog.geometry("500x200")
            dialog.transient(self.root)
            dialog.grab_set()
            
            label = ctk.CTkLabel(dialog, text=message, wraplength=450, font=ctk.CTkFont(size=12))
            label.pack(pady=20, padx=20)
            
            button_frame = ctk.CTkFrame(dialog)
            button_frame.pack(pady=10)
            
            def approve():
                self.approval_response = True
                dialog.destroy()
            
            def deny():
                self.approval_response = False
                dialog.destroy()
            
            ctk.CTkButton(button_frame, text="Allow", command=approve, width=100, fg_color="green").pack(side="left", padx=10)
            ctk.CTkButton(button_frame, text="Deny", command=deny, width=100, fg_color="red").pack(side="left", padx=10)
        else:
            # Fallback for standard tkinter
            import tkinter.messagebox as messagebox
            self.approval_response = messagebox.askyesno("Permission Required", message)

    def _toggle_tts(self):
        """Toggle text-to-speech."""
        if CTK_AVAILABLE:
            self.tts_enabled = self.tts_toggle.get()
        else:
            self.tts_enabled = self.tts_var.get()
        
        status = "enabled" if self.tts_enabled else "disabled"
        self._append_message("System", f"Text-to-speech {status}", "info")

    def _toggle_free_access(self):
        """Toggle free access mode (auto-approve permissions)."""
        try:
            enabled = self.free_access_toggle.get() if CTK_AVAILABLE else self.free_access_var.get()
            status = "enabled" if enabled else "disabled"
            self._append_message("System", f"Free Access {status}", "info")
        except Exception:
            pass

    def _start_voice_listener(self):
        """Start the voice command listener."""
        def on_voice_command(command: str, intent: dict):
            """Handle voice commands."""
            self.root.after(0, lambda: self._append_message("You (voice)", command))
            self.root.after(0, lambda: self._process_message(command))
        
        try:
            self.voice_listener = VoiceListener(
                wake_word="jessica",
                command_callback=on_voice_command
            )
            threading.Thread(target=self.voice_listener.start, daemon=True).start()
        except Exception as e:
            print(f"[UI] Voice listener not started: {e}")
            self._append_message("System", "Voice commands unavailable. Install dependencies: pip install openai-whisper sounddevice soundfile pynput", "info")

    def _attach_files(self):
        """Open file picker to attach files/images."""
        try:
            from tkinter import filedialog
            file_types = [
                ("All Files", "*.*"),
                ("Images", "*.png *.jpg *.jpeg *.gif *.bmp *.webp"),
                ("Documents", "*.pdf *.docx *.doc *.xlsx *.xls *.txt *.csv"),
                ("Text", "*.txt *.md *.log"),
            ]
            files = filedialog.askopenfilenames(
                title="Attach files for analysis",
                filetypes=file_types
            )
            if files:
                self.attached_files = [Path(f) for f in files]
                file_names = ", ".join([f.name for f in self.attached_files])
                self._append_message("System", f"Attached: {file_names}", "info")
        except Exception as e:
            self._append_message("System", f"Error attaching files: {e}", "error")

    def _extract_file_contents(self, file_paths: list) -> str:
        """Extract and parse content from attached files."""
        contents = []
        for file_path in file_paths:
            try:
                suffix = file_path.suffix.lower()
                
                # Handle images
                if suffix in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp']:
                    contents.append(f"[IMAGE: {file_path.name}] Image attached for visual analysis.")
                
                # Handle PDFs
                elif suffix == '.pdf':
                    try:
                        import PyPDF2
                        with open(file_path, 'rb') as f:
                            reader = PyPDF2.PdfReader(f)
                            text = "\n".join([page.extract_text() for page in reader.pages])
                        contents.append(f"[PDF: {file_path.name}]\n{text[:2000]}")  # Limit to 2000 chars
                    except ImportError:
                        contents.append(f"[PDF: {file_path.name}] (PDF library not installed)")
                
                # Handle Word docs
                elif suffix in ['.docx', '.doc']:
                    try:
                        from docx import Document
                        doc = Document(file_path)
                        text = "\n".join([p.text for p in doc.paragraphs])
                        contents.append(f"[WORD: {file_path.name}]\n{text[:2000]}")
                    except ImportError:
                        contents.append(f"[WORD: {file_path.name}] (python-docx library not installed)")
                
                # Handle Excel files
                elif suffix in ['.xlsx', '.xls']:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        ws = wb.active
                        
                        # Get sheet data in a readable format
                        rows = []
                        for row in ws.iter_rows(values_only=True, max_row=100):  # Limit to 100 rows
                            rows.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
                        
                        text = "\n".join(rows[:50])  # Show first 50 rows
                        row_count = ws.max_row
                        col_count = ws.max_column
                        
                        contents.append(f"[EXCEL: {file_path.name}] ({row_count} rows, {col_count} columns)\n{text}")
                    except ImportError:
                        contents.append(f"[EXCEL: {file_path.name}] (openpyxl library not installed)")
                    except Exception as e:
                        contents.append(f"[EXCEL: {file_path.name}] Error reading: {e}")
                
                # Handle text files
                elif suffix in ['.txt', '.md', '.log', '.csv']:
                    text = file_path.read_text(encoding='utf-8', errors='ignore')
                    contents.append(f"[TEXT: {file_path.name}]\n{text[:2000]}")
                
                else:
                    contents.append(f"[FILE: {file_path.name}] Unsupported format.")
            
            except Exception as e:
                contents.append(f"[ERROR reading {file_path.name}]: {e}")
        
        return "\n---\n".join(contents)

    def _send_message(self):
        """Send text message to Jessica."""
        text = self.text_input.get().strip()
        if not text and not self.attached_files:
            return
        
        self.text_input.delete(0, "end")
        
        # Show user message
        if self.attached_files:
            file_names = ", ".join([f.name for f in self.attached_files])
            self._append_message("You", f"{text} [📎 {file_names}]" if text else f"[📎 {file_names}]")
        else:
            self._append_message("You", text)
        
        # Copy attachments before clearing
        attached_files_copy = self.attached_files.copy() if self.attached_files else []
        
        # Clear attachments immediately
        self.attached_files = []
        
        # Process in background thread with attachments
        threading.Thread(target=self._process_message, args=(text, attached_files_copy), daemon=True).start()

    def _show_privacy_settings(self):
        """Open privacy settings dialog with consent toggles."""
        from jessica.automation.consent_manager import ConsentManager
        
        consent_mgr = ConsentManager()
        consents = consent_mgr.get_all()
        
        # Create dialog
        if CTK_AVAILABLE:
            dialog = ctk.CTkToplevel(self.root)
            dialog.title("Privacy & Automation Settings")
            dialog.geometry("500x600")
            
            # Title
            title = ctk.CTkLabel(
                dialog,
                text="Control Jessica's Capabilities",
                font=ctk.CTkFont(size=18, weight="bold")
            )
            title.pack(pady=15)
            
            # Description
            desc = ctk.CTkLabel(
                dialog,
                text="Enable or disable what Jessica can do on your computer.\\nChanges take effect immediately.",
                font=ctk.CTkFont(size=12)
            )
            desc.pack(pady=5)
            
            # Scrollable frame for toggles
            scroll_frame = ctk.CTkScrollableFrame(dialog, width=450, height=400)
            scroll_frame.pack(pady=10, padx=20, fill="both", expand=True)
            
            toggles = {}
            
            capabilities = [
                ("vscode", "VS Code Control", "Create/edit files, run commands in VS Code"),
                ("excel", "Excel Automation", "Open and edit Excel spreadsheets"),
                ("word", "Word Automation", "Open and edit Word documents"),
                ("powerpoint", "PowerPoint Automation", "Control PowerPoint presentations"),
                ("browser", "Browser Control", "Open URLs and perform web searches"),
                ("screen_capture", "Screen Capture", "Take screenshots and see your screen"),
                ("keyboard_mouse", "Keyboard & Mouse", "Type text and control mouse (advanced)"),
            ]
            
            for key, name, desc_text in capabilities:
                frame = ctk.CTkFrame(scroll_frame)
                frame.pack(pady=8, padx=10, fill="x")
                
                label = ctk.CTkLabel(
                    frame,
                    text=f"{name}\\n{desc_text}",
                    font=ctk.CTkFont(size=11),
                    anchor="w",
                    justify="left"
                )
                label.pack(side="left", padx=10, pady=5, fill="x", expand=True)
                
                toggle = ctk.CTkSwitch(
                    frame,
                    text="",
                    onvalue=True,
                    offvalue=False,
                    command=lambda k=key: self._toggle_consent(k, consent_mgr)
                )
                if consents.get(key, False):
                    toggle.select()
                toggle.pack(side="right", padx=10)
                toggles[key] = toggle
            
            # Close button
            close_btn = ctk.CTkButton(
                dialog,
                text="Close",
                command=dialog.destroy,
                width=120
            )
            close_btn.pack(pady=15)
            
        else:
            # Fallback for vanilla Tkinter
            dialog = tk.Toplevel(self.root)
            dialog.title("Privacy & Automation Settings")
            dialog.geometry("500x600")
            
            tk.Label(
                dialog,
                text="Privacy & Automation Settings",
                font=("Arial", 16, "bold")
            ).pack(pady=10)
            
            tk.Label(
                dialog,
                text="Enable or disable Jessica's capabilities",
                font=("Arial", 10)
            ).pack(pady=5)
            
            text = tk.Text(dialog, width=60, height=20)
            text.pack(pady=10)
            text.insert("1.0", "Privacy settings available with CustomTkinter.\\n\\n" + str(consents))
            text.config(state="disabled")
            
            tk.Button(dialog, text="Close", command=dialog.destroy).pack(pady=10)

    def _toggle_consent(self, capability: str, consent_mgr):
        """Toggle a consent setting."""
        if consent_mgr.is_allowed(capability):
            consent_mgr.revoke(capability)
            self._append_message("System", f"❌ {capability.replace('_', ' ').title()} disabled", "info")
        else:
            consent_mgr.grant(capability)
            self._append_message("System", f"✅ {capability.replace('_', ' ').title()} enabled", "info")

    def run(self):
        """Start the UI main loop."""
        print("[UI] Starting Jessica Desktop UI...")
        self.root.mainloop()


def main():
    """Launch the desktop UI."""
    if not CTK_AVAILABLE:
        print("Warning: CustomTkinter not installed. Using standard Tkinter.")
        print("For better UI, install: pip install customtkinter")
    
    ui = JessicaDesktopUI()
    ui.run()


if __name__ == "__main__":
    main()
